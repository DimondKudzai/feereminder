import os, csv, time
from pathlib import Path
from datetime import datetime
from functools import wraps

import fitz
import openpyxl
import requests
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from sqlalchemy import func, case
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv

load_dotenv()

# ====================== APP & DB SETUP ======================


app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', '2409')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Optional but recommended for Render/Neon
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_pre_ping": True,       # checks connections are alive
    "pool_recycle": 280,         # recycle idle connections before timeout
    "connect_args": {"sslmode": "require"}  # ensures SSL
}


db = SQLAlchemy(app)

# ====================== MODELS ======================
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.Text, unique=True, nullable=False)
    password_hash = db.Column(db.Text, nullable=False)
    school_name = db.Column(db.Text, nullable=False)
    sender_id = db.Column(db.Text, nullable=False)
    paycode = db.Column(db.Text)
    sms_credits = db.Column(db.Integer, default=1000)
    plan = db.Column(db.Text, default='pro')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    messages = db.relationship('Message', backref='user', lazy=True)
    tickets = db.relationship('Ticket', backref='user', lazy=True)

class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    student_name = db.Column(db.Text)
    phone = db.Column(db.Text)
    message = db.Column(db.Text)
    msg_id = db.Column(db.Text)
    status = db.Column(db.Text, default='queued')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Ticket(db.Model):
    __tablename__ = 'tickets'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    subject = db.Column(db.Text)
    body = db.Column(db.Text)
    status = db.Column(db.Text, default='open')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

with app.app_context():
    db.create_all()
    
# ====================== SMS SENDER ======================
def send_sms_sync(user_id, recipients):
    import os, requests, time
 #  from your_models import db, Message, User  # adjust imports

    api_key = os.getenv('PING_API_KEY')
    if not api_key:
        raise ValueError("PING_API_KEY not set")

    url = "https://api.ping.co.zw/v1/notification/api/sms/send"
    headers = {"X-Ping-Api-Key": api_key, "Content-Type": "application/json"}

    success_count = 0
    failed_count = 0

    for recipient in recipients:
        phone = ''.join(filter(str.isdigit, str(recipient['phone'])))
        if phone.startswith('0'):
            phone = '+263' + phone[1:]
        elif not phone.startswith('+263'):
            phone = '+263' + phone

        # New message template with dynamic fields
        msg = (
            f"Good day {recipient['name']}, We hope you are well.\n"
            f"This is our new platform where we remind you about your monthly fee payments. "
            f"Thank you for continuous support\n"
            f"##Building strong minds for a brighter future."
        )
        
        payload = {"to_phone": phone, "message": msg}

        try:
            r = requests.post(url, headers=headers, json=payload, timeout=10)
            resp = r.json()
            success = r.status_code == 200 and resp.get('status') == 'success'
            msg_id = resp.get('messageId', '')
            status = 'sent' if success else 'failed'

            if success:
                success_count += 1
            else:
                failed_count += 1

        except Exception:
            msg_id, status = '', 'failed'
            failed_count += 1

        db.session.add(Message(
            user_id=user_id,
            student_name=recipient['name'],
            phone=phone,
            message=msg,
            msg_id=msg_id,
            status=status
        ))
        db.session.commit()
        time.sleep(0.2)

    if success_count > 0:
        user = User.query.get(user_id)
        user.sms_credits = max(0, user.sms_credits - success_count)
        db.session.commit()

    return {"sent": success_count, "failed": failed_count, "total": len(recipients)}
# ====================== FILE PARSING ======================
def parse_pdf(filepath):
    doc = fitz.open(filepath)
    rows, headers = [], []
    for page in doc:
        tables = page.find_tables()
        if tables.tables:
            data = tables.tables[0].extract()
            if not data:
                continue
            headers = [str(h or '').lower().strip() for h in data[0]]
            for row in data[1:]:
                if any(cell and str(cell).strip() for cell in row):
                    clean_row = [str(c or '').strip() for c in row]
                    rows.append(dict(zip(headers, clean_row)))
            break
    doc.close()
    if not headers:
        raise ValueError("No table found in PDF. Export as Excel if this fails.")
    return rows, headers

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(cell.value or '').lower().strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [str(cell.value or '').strip() for cell in row]
        if any(values):
            rows.append(dict(zip(headers, values)))
    return rows, headers

def parse_csv(filepath):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = [h.lower().strip() for h in reader.fieldnames]
        rows = [{k.lower().strip(): str(v or '').strip() for k, v in row.items()} for row in reader]
    return rows, headers

def parse_any_file(filepath):
    ext = Path(filepath).suffix.lower()
    if ext == '.pdf':
        return parse_pdf(filepath)
    elif ext in ['.xlsx', '.xls']:
        return parse_excel(filepath)
    elif ext == '.csv':
        return parse_csv(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use PDF, Excel, or CSV.")

def auto_find_cols(headers):
    name_keys = ['name', 'student', 'learner', 'pupil']
    phone_keys = ['phone', 'mobile', 'cell', 'contact', 'number']
    bal_keys = ['balance', 'due', 'owing', 'amount', 'fee', 'outstanding']

    name_c = next((h for h in headers if any(k in h for k in name_keys)), None)
    phone_c = next((h for h in headers if any(k in h for k in phone_keys)), None)
    bal_c = next((h for h in headers if any(k in h for k in bal_keys)), None)

    if not all([name_c, phone_c, bal_c]):
        raise ValueError(f"Missing columns. Found: {headers}. Need name, phone, balance columns.")
    return name_c, phone_c, bal_c

# ====================== AUTH DECORATORS ======================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.context_processor
def inject_user():
    if 'user_id' in session:
        return dict(current_user=User.query.get(session['user_id']))
    return dict(current_user=None)

# ====================== WEBHOOK ======================
@app.route('/webhook', methods=['POST'])
def ping_webhook():
    data = request.get_json()
    msg_id = data.get('messageId')
    status = data.get('status')
    if msg_id:
        Message.query.filter_by(msg_id=msg_id).update({'status': status})
        db.session.commit()
    return '', 200

# ====================== PWA ======================
@app.route('/manifest.json')
def manifest():
    return jsonify({
        "name": "FeeRemind Pro",
        "short_name": "FeeRemind",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#0f172a",
        "theme_color": "#2563eb",
        "icons": [{"src": "/static/ic-1.png", "sizes": "192x192"}]
    })

@app.route('/sw.js')
def sw():
    return app.send_static_file('sw.js')

# ====================== AUTH ROUTES ======================
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        school_name = request.form['school_name']
        sender_id = request.form['sender_id']
        paycode = request.form['paycode']

        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'error')
            return render_template('register.html')

        user = User(
            email=email,
            password_hash=generate_password_hash(password),
            school_name=school_name,
            sender_id=sender_id,
            paycode=paycode
        )
        db.session.add(user)
        db.session.commit()
        flash('Account created. Login now.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            return redirect(url_for('dashboard'))
        flash('Invalid credentials', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ====================== APP ROUTES ======================
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))



@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    today = datetime.utcnow().date()

    # ====== Simplified Stats ======
    stats = db.session.query(
        func.count(Message.id).label('total'),
        func.count(case((Message.status == 'Delivered', 1))).label('delivered'),
        func.count(case((Message.status == 'Failed', 1))).label('failed'),
        func.count(case((Message.status == 'sent', 1))).label('sent')
    ).filter(
        Message.user_id == user.id,
        func.date(Message.created_at) == today
    ).first()

    # ====== Recent Uploads ======
    uploads = db.session.query(
        func.date(Message.created_at).label('date'),
        func.count(Message.id).label('count'),
        func.count(case((Message.status == 'Delivered', 1))).label('delivered')
    ).filter(
        Message.user_id == user.id
    ).group_by(
        func.date(Message.created_at)
    ).order_by(
        func.date(Message.created_at).desc()
    ).limit(7).all()

    return render_template(
        'dashboard.html',
        user=user,
        stats=stats,
        uploads=uploads,
        school_name=user.school_name
    )
    
    
@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    user = User.query.get(session['user_id'])
    if user.sms_credits <= 0:
        flash('SMS credits exhausted. Contact support.', 'error')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        file = request.files['file']
        if not file.filename:
            flash('No file selected', 'error')
            return redirect(url_for('upload'))

        path = f"/tmp/{file.filename}"
        file.save(path)

        try:
            rows, headers = parse_any_file(path)
            name_c, phone_c, bal_c = auto_find_cols(headers)
        except Exception as e:
            flash(f'File error: {str(e)}', 'error')
            os.remove(path)
            return redirect(url_for('upload'))

        recipients = []
        for r in rows:
            try:
                bal = float(str(r.get(bal_c, '0')).replace('$','').replace(',','') or 0)
            except:
                continue
            if bal: # > 0 and user.sms_credits > len(recipients):
                recipients.append({
                    "name": str(r.get(name_c, 'Parent')),
                    "phone": str(r.get(phone_c, '')),
                    "balance": bal,
                    "due": str(r.get('due', r.get('due_date', r.get('duedate', 'ASAP')))),
                    "paycode": user.paycode or ''
                })

        os.remove(path)

        # ADD THE 100 ROW LIMIT
        MAX_ROWS_PER_UPLOAD = 100
        if len(recipients) > MAX_ROWS_PER_UPLOAD:
            flash(f'Too many valid rows: {len(recipients)}. Max allowed is {MAX_ROWS_PER_UPLOAD}. Split your file and try again.', 'error')
            return redirect(url_for('upload'))

        if recipients:
            result = send_sms_sync(user.id, recipients)
            flash(f"Sent {result['sent']}, failed {result['failed']} out of {result['total']}", 'success')
        else:
            flash('No valid recipients found', 'error')

        return redirect(url_for('dashboard'))

    return render_template('upload.html', user=user, school_name=user.school_name)

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    user = User.query.get(session['user_id'])
    if request.method == 'POST':
        user.school_name = request.form['school_name']
        user.sender_id = request.form['sender_id']
        user.paycode = request.form['paycode']
        db.session.commit()
        flash('Settings saved', 'success')
        return redirect(url_for('settings'))
    return render_template('settings.html', user=user, school_name=user.school_name)

@app.route('/api/status')
@login_required
def api_status():
    user = User.query.get(session['user_id'])
    today = datetime.utcnow().date()
    data = dict(db.session.query(Message.status, db.func.count(Message.id))
               .filter(Message.user_id == user.id, db.func.date(Message.created_at) == today)
               .group_by(Message.status).all())
    return jsonify(data)

@app.route('/help', methods=['GET', 'POST'])
@login_required
def help():
    user = User.query.get(session['user_id'])
    if request.method == 'POST':
        ticket = Ticket(user_id=user.id, subject=request.form['subject'], body=request.form['body'])
        db.session.add(ticket)
        db.session.commit()
        flash('Ticket submitted. We reply within 4 hours.', 'success')
    return render_template('help.html', user=user, school_name=user.school_name)

# ====================== RUN APP ======================
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))